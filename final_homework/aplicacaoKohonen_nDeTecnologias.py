from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

# book = load_workbook('results_teste.xlsx')
book = load_workbook('results_pleno_completo.xlsx')
sheet = book.active
from trabalho_final.kohonen import kohonen

i = 2
end = False

jobs_dict = dict()

while not end:
    if not sheet['A' + str(i)].value == None:
        # jobs_dict['A' + str(i)] = [int(sheet['A' + str(i)].value[2:].replace('.', '')), []]
        jobs_dict['A' + str(i)] = [float(sheet['A' + str(i)].value[2:].replace('.', '')) / 1000, []]
        j = 1

        while not sheet[chr(ord(chr(ord('B') + j))) + str(i)].value == None:
            jobs_dict['A' + str(i)][1].append(sheet[chr(ord(chr(ord('B') + j))) + str(i)].value)
            j += 1

        i += 1
    else:
        end = True


initial_SIGMA = 1
initial_ALPHA = 0.2

n_epocas = 1

SIGMA_array = np.linspace(initial_SIGMA, 0.1, n_epocas)
ALPHA_array = np.linspace(initial_ALPHA, 0.1, n_epocas)

w_arrays = np.zeros((1000, 2))

for index, array in enumerate(w_arrays):
    # print(index, array)
    # w_arrays[index] = [np.random.randint(1000, 10000), np.random.random() * 2]
    w_arrays[index] = [np.random.random() * 10, np.random.random() * 10]


x = [array[1] for array in w_arrays]
y = [array[0] for array in w_arrays]
# plt.scatter(x=x, y=y, label='Pesos iniciais', color='black')
plt.scatter(x=x, y=y, label='Initial weights', color='gray', edgecolors='black', marker='_')
plt.legend()

plt.title('number of epochs: ' + str(n_epocas))
plt.xlabel('number of required technologies')
plt.ylabel('Salary (x2000)')
plt.savefig('nTechnologies_initial')
plt.show()
plt.close()

entry = np.array([])

current_job = 0
for epoca in range(n_epocas):
    print('epoca: ', epoca)
    for key, value in jobs_dict.items():
        print(key, jobs_dict.__len__())
        # print('\n' + '-' * 30)
        # print(key, value)
        temp_atributes = []

        entry = np.array([jobs_dict[key][0], value[1].__len__()])

        # w_arrays = kohonen(w_arrays=w_arrays, entry=entry, SIGMA=initial_SIGMA, ALPHA=initial_ALPHA, n_epocas=1)
        w_arrays = kohonen(w_arrays=w_arrays, entry=entry, SIGMA=SIGMA_array[epoca], ALPHA=ALPHA_array[epoca], n_epocas=1)


        current_job += 1

    current_job = 0

print(w_arrays)
x = [array[1] for array in w_arrays]
y = [array[0] for array in w_arrays]
plt.scatter(x=x, y=y, label='final weights', color='gray', edgecolors='black', marker='_')
plt.legend()

plt.title('number of epochs: ' + str(n_epocas))
plt.xlabel('number of required technologies')
plt.ylabel('Salary (x2000)')

plt.savefig('ntechnologies_final.png')
plt.show()
