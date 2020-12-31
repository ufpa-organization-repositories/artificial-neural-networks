from trabalho_final.backpropagation_final import train_backpropagation, apply_backpropagation, plotErrors_comparative, plotErrors
from openpyxl import load_workbook
import numpy as np

def get_jobs_dict_from_xls_file(sheet):
    i = 2
    end = False

    jobs_dict = dict()
    while not end:
        if not sheet['A' + str(i)].value == None:
            jobs_dict['A' + str(i)] = [int(sheet['A' + str(i)].value[2:].replace('.', '')), []]
            j = 2

            while not sheet[chr(ord(chr(ord('B') + j))) + str(i)].value == None:
                jobs_dict['A' + str(i)][1].append(sheet[chr(ord(chr(ord('B') + j))) + str(i)].value)
                j += 1

            i += 1
        else:
            end = True

    return jobs_dict

def get_tecnologies_from_jobs_dict(jobs_dict):
    tecnologies = dict()
    for args in jobs_dict.values():
        # print(args[1][2:])

        for tecnology in args[1][2:]:


            try:

                if type(tecnologies[tecnology]) == int:
                    tecnologies[tecnology] += 1

                elif tecnologies[tecnology] == tecnology:
                    tecnologies[tecnology] = 1

            except:

                tecnologies[tecnology] = tecnology


    return tecnologies

def get_especfic_entry_tecnologies(jobs_dict):
    entradas = []
    saidas = []

    for key, value in jobs_dict.items():
        entrada = {'JavaScript': 0,
                   'CSS': 0,
                   'React Native': 0,
                   'Node.js': 0,
                   'ReactJS': 0,
                   'PHP': 0,
                   'HTML': 0,
                   'Angular': 0,
                   'Vue.js': 0,
                   'Laravel': 0,
                   'MySQL': 0,
                   'Ionic': 0,
                   '.NET': 0,
                   'SQL Server': 0,
                   'AWS RDS (Relational Database Service)': 0,
                   'AWS S3': 0,
                   'C#': 0,
                   'MongoDB': 0}

        # print('\n' + '-' * 30)
        # print(key, value)

        for tec in value[1][2:]:
            if tec in entrada:
                entrada[tec] = 1

        # saida = value[0] / 20000
        # saida = [0, 0, 0, 0, 0, 0]
        saida = [0, 0]

        if value[0] < 5000:
            saida[0] = 1
            print(saida)


        else:
            saida[1] = 1
            print(saida)

        # print(saida, entrada)

        entradas.append(entrada)
        saidas.append(saida)

    return entradas, saidas

def ajust_entries_and_exits(entradas, saidas):
    entradas_final = []
    saidas_final = []

    for entrada, saida in zip(entradas, saidas):
        # print(np.array(list(entrada.keys())))
        # print(np.array(list(entrada.values())))

        entrada_temp = list(entrada.values())
        # print(entrada_temp, saida)
        # entradas_temp.reshape(entradas_temp.__len__(), 1)

        entradas_final.append(entrada_temp)
        # saidas_final.append(saida)

    entradas_final = np.asarray(entradas_final)
    # saidas_final = np.asarray(saidas).reshape(len(saidas), 1)
    saidas_final = np.asarray(saidas)
    for e, s in zip(entradas_final, saidas_final):
        print(e, s)

    return entradas_final, saidas_final

# book = load_workbook('results_teste.xlsx')
# book = load_workbook('pleno_40_primeiros.xlsx')
book = load_workbook('results_São Paulocltstartuppleno.xlsx')
sheet = book.active

jobs_dict = get_jobs_dict_from_xls_file(sheet=sheet)
tecnologies = get_tecnologies_from_jobs_dict(jobs_dict=jobs_dict)
entradas, saidas = get_especfic_entry_tecnologies(jobs_dict=jobs_dict)
entradas, saidas = ajust_entries_and_exits(entradas=entradas, saidas=saidas)

n_neuronios_camadaOculta = 3
epocas =100000
taxaAprendizagem = 0.0001
momento = 1

n_entradas = entradas[0].__len__()
n_saidas = saidas[0].__len__()

# pesos0 = 2 * np.random.random((n_entradas, n_neuronios_camadaOculta)) - 1
# pesos1 = 2 * np.random.random((n_neuronios_camadaOculta, n_saidas)) - 1

pesos0 = 2 * np.random.random((n_entradas, n_neuronios_camadaOculta)) - 1
pesos1 = 2 * np.random.random((n_neuronios_camadaOculta, n_saidas)) - 1

pesos0, pesos1, camadaSaida, error_array = train_backpropagation(epocas=epocas,
                                                                 entradas=entradas, pesos0=pesos0,
                                                                 saidas=saidas, pesos1=pesos1,
                                                                 taxaAprendizagem=taxaAprendizagem,
                                                                 momento=momento)

book = load_workbook('pleno_40_primeiros.xlsx')
# book = load_workbook('results_São Paulocltstartupjunior.xlsx')
sheet = book.active

jobs_dict = get_jobs_dict_from_xls_file(sheet=sheet)
tecnologies = get_tecnologies_from_jobs_dict(jobs_dict=jobs_dict)
entradas, saidas = get_especfic_entry_tecnologies(jobs_dict=jobs_dict)
entradas, saidas = ajust_entries_and_exits(entradas=entradas, saidas=saidas)

camadaSaida = apply_backpropagation(entradas=entradas, pesos0=pesos0, pesos1=pesos1, error_array=error_array)
plotErrors_comparative(camadaSaida, saidas)
print(pesos0)
print(pesos1)







